import openpyxl

# 讀取最新嘅Excel文件
file_path = '/data/.openclaw/media/inbound/ACS_Order_APR2026_CS---997ab6ed-a3ce-4c12-bc31-052ab333e6c6.xlsx'
wb = openpyxl.load_workbook(file_path)

print('=== 處理訂單546271（11月3日）===')

# 檢查是否已有3月嘅sheet
if '11MAR2026' in wb.sheetnames:
    print('11MAR2026 sheet 已存在')
    ws = wb['11MAR2026']
else:
    print('創建11MAR2026 sheet')
    ws = wb.create_sheet('11MAR2026')
    
    # 設置標題行
    headers = ['Ord No.', 'Time', 'Trip', 'Driver', 'Status', 'Pick', 'Drop', '', '', 'Driver.Details', 'Dvr.OT', 'Dvr.C', 'Basic', 'Dvr.F', 'Bonus']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

# 檢查現有訂單
existing_orders = []
for row in range(2, ws.max_row + 1):
    order_no = ws.cell(row=row, column=1).value
    if order_no and order_no != 'Ord No.' and str(order_no).strip():
        try:
            existing_orders.append(float(order_no))
        except:
            pass

print(f'現有訂單: {existing_orders}')

# 訂單546271嘅數據
order_to_check = 546271
if order_to_check in existing_orders:
    print(f'訂單 {order_to_check} 已存在，跳過')
else:
    print(f'訂單 {order_to_check} 不存在，準備輸入')
    
    # 尋找第一個空行
    start_row = 2
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value is None or cell_value == '':
            start_row = row
            break
    
    print(f'從第{start_row}行開始輸入訂單')
    
    # 計算各項費用
    ot_fee = 400  # 大嶼山附加費
    parking_fee = 16  # 停車場費
    tunnel_fee = 16  # 隧道費
    
    # 總附加費 = 大嶼山附加費 + 停車場費
    total_extra = ot_fee + parking_fee
    
    # 司機費 = 基本價$400 + 附加費
    driver_fee = 400 + total_extra
    
    # 訂單546271嘅數據
    order_data = {
        'ord_no': 546271,
        'time': '12:06',
        'trip': 'Ap - yau ma tei - K11 - che kung - Disney',
        'driver': 'Keith',
        'plate': '',  # 原始資料冇提供車牌
        'mobile': '',  # 原始資料冇提供電話
        'pick': '12:06',
        'drop': '16:52',
        'extra_fees': f'大嶼山附加費2X200=${ot_fee}, 停車场费${parking_fee}, 隧道费${tunnel_fee}',
        'ot': ot_fee,
        'parking': parking_fee,
        'basic_price': 400,
        'driver_fee': driver_fee,
        'bonus': 30
    }
    
    # 輸入訂單數據
    row_num = start_row
    
    # 設置單元格數據
    ws.cell(row=row_num, column=1, value=order_data['ord_no'])  # Ord No.
    ws.cell(row=row_num, column=2, value=order_data['time'])    # Time
    ws.cell(row=row_num, column=3, value=order_data['trip'])    # Trip
    ws.cell(row=row_num, column=4, value=f'{order_data["driver"]} {order_data["plate"]} {order_data["mobile"]}')  # Driver + Plate + Mobile
    ws.cell(row=row_num, column=5, value='Done')                # Status
    ws.cell(row_num, column=6, value=order_data['pick'])    # Pick
    ws.cell(row_num, column=7, value=order_data['drop'])    # Drop
    ws.cell(row_num, column=8, value=order_data['extra_fees'])  # Extra fees
    ws.cell(row_num, column=9, value=order_data['pick'])    # Ptime
    ws.cell(row_num, column=10, value=f'{order_data["driver"]} {order_data["plate"]} {order_data["mobile"]}')  # Driver details
    ws.cell(row_num, column=11, value=order_data['ot'])     # Dvr.OT
    ws.cell(row_num, column=12, value=order_data['parking']) # Dvr.C
    ws.cell(row_num, column=13, value=order_data['basic_price']) # Basic
    ws.cell(row_num, column=14, value=order_data['driver_fee']) # Dvr.F
    ws.cell(row_num, column=15, value=order_data['bonus'])      # Bonus
    
    print(f'已輸入訂單 {order_data["ord_no"]} 到第{row_num}行')
    print(f'司機: {order_data["driver"]}')
    print(f'路線: {order_data["trip"]}')
    print(f'時間: {order_data["pick"]} - {order_data["drop"]} (5小時)')
    print(f'附加費: 大嶼山附加費${ot_fee}, 停車场费${parking_fee}')
    print(f'司機費: ${driver_fee}')
    print(f'總費用: ${driver_fee + order_data["bonus"]}')

# 保存修改後嘅文件
output_path = '/data/workspace/ACS_Order_APR2026_Order546271.xlsx'
wb.save(output_path)
print(f'\n📁 文件已保存到: {output_path}')
print('✅ 訂單546271處理完成！')