import openpyxl
from openpyxl.styles import Font, Alignment
import os

def input_orders_to_excel():
    try:
        file_path = '/data/.openclaw/media/inbound/ACS_Order_APR2026_CS---3f1c2ee7-7992-4675-a8a5-3f1c156e2618.xlsx'
        
        # 加載工作簿
        wb = openpyxl.load_workbook(file_path)
        ws = wb['1APR2026']
        
        print(f'當前工作表大小: {ws.max_row}行 x {ws.max_column}列')
        
        # 訂單數據
        orders = [
            {
                'ord_no': '549294',
                'time': '3:00',
                'trip': 'TST > AP',
                'driver': 'Eva',
                'plate': 'TZ9800',
                'mobile': '97062468',
                'pick': '2:59',
                'drop': '3:32',
                'decp': 'OT$60',
                'ptime': '2:59',
                'driver_details': 'Eva TZ9800 97062468',
                'ot': '60',
                'status': ''
            },
            {
                'ord_no': '549296',
                'time': '4:30',
                'trip': 'TST > AP',
                'driver': 'Jacky',
                'plate': 'CT8022',
                'mobile': '69986953',
                'pick': '4:29',
                'drop': '5:00',
                'decp': 'OT$60',
                'ptime': '4:29',
                'driver_details': 'Jacky CT8022 69986953',
                'ot': '60',
                'status': ''
            },
            {
                'ord_no': '547719',
                'time': '11:45',
                'trip': 'CX983 > 迪士尼荷里活酒店',
                'driver': 'Keith 陳',
                'plate': 'CK9898',
                'mobile': '98880261',
                'pick': '12:05',
                'drop': '12:25',
                'decp': '2 booster',
                'ptime': '12:05',
                'driver_details': 'Keith 陳 CK9898 98880261',
                'ot': '',
                'status': ''
            },
            {
                'ord_no': '549376',
                'time': '10:25 AM',
                'trip': 'Cordis > AP',
                'driver': 'Keith 陳',
                'plate': 'CK9898',
                'mobile': '98880261',
                'pick': '10:28',
                'drop': '11:00',
                'decp': '',
                'ptime': '10:28',
                'driver_details': 'Keith 陳 CK9898 98880261',
                'ot': '',
                'status': ''
            },
            {
                'ord_no': '549375',
                'time': '09:00 AM',
                'trip': '3V 401 > AP',
                'driver': 'Eva',
                'plate': 'TZ9800',
                'mobile': '97062468',
                'pick': '9:08',
                'drop': '10:01',
                'decp': '',
                'ptime': '9:08',
                'driver_details': 'Eva TZ9800 97062468',
                'ot': '',
                'status': ''
            },
            {
                'ord_no': '549375',
                'time': '09:00 AM',
                'trip': '3V 401 > AP',
                'driver': 'Cheuk/卓師傅',
                'plate': 'KC6089',
                'mobile': '96695358',
                'pick': '9:08',
                'drop': '10:01',
                'decp': 'P$35',
                'ptime': '9:08',
                'driver_details': 'Cheuk/卓師傅 KC6089 96695358',
                'ot': '',
                'status': ''
            },
            {
                'ord_no': '549295',
                'time': '4:00',
                'trip': 'MB294 > AP',
                'driver': 'Eva',
                'plate': 'TZ9800',
                'mobile': '97062468',
                'pick': '4:38',
                'drop': '5:05',
                'decp': '',
                'ptime': '4:38',
                'driver_details': 'Eva TZ9800 97062468',
                'ot': '60',
                'status': ''
            },
            {
                'ord_no': '549295',
                'time': '4:00',
                'trip': 'MB294 > AP',
                'driver': 'Ricky 黎',
                'plate': 'ZJ6071',
                'mobile': '51089715',
                'pick': '4:38',
                'drop': '5:05',
                'decp': 'P$35',
                'ptime': '4:38',
                'driver_details': 'Ricky 黎 ZJ6071 51089715',
                'ot': '60',
                'status': ''
            }
        ]
        
        # 尋找第一個空行
        start_row = 2
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value is None or cell_value == '':
                start_row = row
                break
        
        print(f'從第{start_row}行開始輸入訂單')
        
        # 輸入訂單數據
        for i, order in enumerate(orders):
            row_num = start_row + i
            
            # 設置單元格數據
            ws.cell(row=row_num, column=1, value=order['ord_no'])  # Ord No.
            ws.cell(row=row_num, column=2, value=order['time'])    # Time
            ws.cell(row=row_num, column=3, value=order['trip'])    # Trip
            ws.cell(row=row_num, column=4, value=order['driver'])  # Driver
            ws.cell(row=row_num, column=5, value=order['status'])  # Status
            ws.cell(row=row_num, column=6, value=order['pick'])    # Pick
            ws.cell(row=row_num, column=7, value=order['drop'])    # Drop
            ws.cell(row=row_num, column=8, value=order['decp'])    # Decp
            ws.cell(row=row_num, column=9, value=order['ptime'])   # Ptime
            ws.cell(row=row_num, column=10, value=order['driver_details'])  # Driver.Details
            
            # 設置 OT 數值
            if order['ot']:
                try:
                    ws.cell(row=row_num, column=11, value=float(order['ot']))  # Dvr.OT
                except ValueError:
                    ws.cell(row=row_num, column=11, value=0)  # Dvr.OT
            
            # 設置 Driver.C（車位費）
            if 'P$' in order.get('decp', ''):
                try:
                    parking_fee = float(order['decp'].replace('P$', ''))
                    ws.cell(row=row_num, column=12, value=parking_fee)  # Dvr.C
                except ValueError:
                    ws.cell(row=row_num, column=12, value=0)  # Dvr.C
            
            print(f'已輸入第{i+1}筆訂單到第{row_num}行: {order["ord_no"]} - {order["trip"]}')
        
        # 保存修改後的文件
        output_path = '/data/workspace/ACS_Order_APR2026_Orders_Input.xlsx'
        wb.save(output_path)
        print(f'文件已保存到: {output_path}')
        
        # 返回輸入的訂單數量
        return len(orders)
        
    except Exception as e:
        print(f'處理文件時出錯: {e}')
        import traceback
        traceback.print_exc()
        return 0

if __name__ == '__main__':
    orders_count = input_orders_to_excel()
    print(f'成功輸入 {orders_count} 筆訂單到 1APR2026 工作表')