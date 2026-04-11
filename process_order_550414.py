import openpyxl

# 讀取最新嘅Excel文件
file_path = 'ACS_Order_APR2026_CS---997ab6ed-a3ce-4c12-bc31-052ab333e6c6.xlsx'
wb = openpyxl.load_workbook(file_path)

print('=== 處理訂單550414（4月5日）===')

# 檢查5月5日sheet
if '5APR2026' in wb.sheetnames:
    print('5APR2026 sheet 已存在')
    ws = wb['5APR2026']
else:
    print('創建5APR2026 sheet')
    ws = wb.create_sheet('5APR2026')
    
    # 設置標題行
    headers = ['Ord No.', 'Time', 'Trip', 'Driver', 'Status', 'Pick', 'Drop', '', '', 'Driver.Details', 'Dvr.OT', 'Dvr.C', 'Basic', 'Dvr.F', 'Bonus']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

# 檢查現有訂單
existing_orders = []
for row in range(2, ws.max_row + 1):
    order_no = ws.cell(row=row, column=1).value
    if order_no and order_no != 'Ord No.' and str(order_no).strip():
        try:
            existing_orders.append(float(order_no))
        except:
            pass

print(f'現有訂單: {existing_orders}')

# 訂單550414嘅數據
order_to_check = 550414
if order_to_check in existing_orders:
    print(f'訂單 {order_to_check} 已存在，跳過')
else:
    print(f'訂單 {order_to_check} 不存在，準備輸入')
    
    # 尋找第一個空行
    start_row = 2
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value is None or cell_value == '':
            start_row = row
            break
    
    print(f'從第{start_row}行開始輸入訂單')
    
    # 訂單550414嘅數據（兩台車）
    order_data = {
        'ord_no': 550414,
        'time': '20:55',
        'trip': 'CX581 > Repulse Bay Gardens',
        'cars': [
            {
                'driver': 'Eva',
                'plate': 'TZ9800',
                'mobile': '97062468',
                'customer_name': 'Ricky 黎',
                'customer_plate': 'ZJ6071',
                'customer_mobile': '51089715',
                'pick': '22:01',
                'drop': '22:50',
                'extra_fees': '南區$60, Bb seat x2',
                'ot': 0,  # 南區費可能不屬於OT
                'parking': 0,
                'bb_seat': 720,  # 兩個兒童座椅
                'basic_price': 400,
                'driver_fee': 1120,  # $400 + $60 + $720
                'bonus': 30
            },
            {
                'driver': 'Ricky 黎',
                'plate': 'ZJ6071',
                'mobile': '51089715',
                'customer_name': '',
                'customer_plate': '',
                'customer_mobile': '',
                'pick': '22:01',
                'drop': '22:50',
                'extra_fees': '南區$60, Bb seat x1',
                'ot': 0,
                'parking': 0,
                'bb_seat': 640,  # 一個booster
                'basic_price': 400,
                'driver_fee': 1100,  # $400 + $60 + $640
                'bonus': 30
            }
        ]
    }
    
    # 輸入訂單數據（每台車一行）
    for i, car in enumerate(order_data['cars']):
        row_num = start_row + i
        
        # 設置單元格數據
        ws.cell(row=row_num, column=1, value=order_data['ord_no'])  # Ord No.
        ws.cell(row=row_num, column=2, value=order_data['time'])    # Time
        ws.cell(row=row_num, column=3, value=order_data['trip'])    # Trip
        ws.cell(row=row_num, column=4, value=f'{car["driver"]} {car["plate"]} {car["mobile"]}')  # Driver + Plate + Mobile
        ws.cell(row=row_num, column=5, value='Done')                # Status
        ws.cell(row=row_num, column=6, value=car['pick'])    # Pick
        ws.cell(row=row_num, column=7, value=car['drop'])    # Drop
        ws.cell(row=row_num, column=8, value=car['extra_fees'])  # Extra fees
        ws.cell(row=row_num, column=9, value=car['pick'])    # Ptime
        ws.cell(row=row_num, column=10, value=f'{car["driver"]} {car["plate"]} {car["mobile"]}')  # Driver details
        
        # 設置OT和車位費（兒童座椅可能需要特殊處理）
        ws.cell(row=row_num, column=11, value=car['ot'])     # Dvr.OT
        ws.cell(row=row_num, column=12, value=car['parking']) # Dvr.C
        
        # 設置基本價、司機費、獎金
        ws.cell(row=row_num, column=13, value=car['basic_price'])    # Basic
        ws.cell(row=row_num, column=14, value=car['driver_fee'])    # Dvr.F
        ws.cell(row=row_num, column=15, value=car['bonus'])         # Bonus
        
        print(f'已輸入第{i+1}台車到第{row_num}行: {car["driver"]} {car["plate"]}')
    
    print(f'\n📋 訂單 {order_data["ord_no"]} 詳細信息:')
    print(f'時間: {order_data["time"]} - 上車: {order_data["cars"][0]["pick"]}, 下車: {order_data["cars"][0]["drop"]}')
    print(f'路線: {order_data["trip"]}')
    print(f'特殊要求: 一台車兩個兒童座椅（5歲）+ 一台車一個booster（7歲）')
    print(f'\n💰 費用計算:')
    for i, car in enumerate(order_data['cars']):
        print(f'第{i+1}台車 ({car["driver"]}): 基本價${car["basic_price"]} + 兒童座椅${car["bb_seat"]} = 司機費${car["driver_fee"]} + 獎金${car["bonus"]}')

# 保存修改後嘅文件
output_path = '/data/workspace/ACS_Order_APR2026_Order550414.xlsx'
wb.save(output_path)
print(f'\n📁 文件已保存到: {output_path}')
print('✅ 訂單550414處理完成！')