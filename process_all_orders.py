import openpyxl
import os

# 讀取最新嘅Excel文件
file_path = 'ACS_Order_APR2026_CS---997ab6ed-a3ce-4c12-bc31-052ab333e6c6.xlsx'
wb = openpyxl.load_workbook(file_path)

print('=== 處理新訂單 ===')

# 處理訂單548798（4月5日）
print('\n=== 處理訂單548798（4月5日）===')
if '5APR2026' in wb.sheetnames:
    ws = wb['5APR2026']
else:
    ws = wb.create_sheet('5APR2026')
    headers = ['Ord No.', 'Time', 'Trip', 'Driver', 'Status', 'Pick', 'Drop', '', '', 'Driver.Details', 'Dvr.OT', 'Dvr.C', 'Basic', 'Dvr.F', 'Bonus']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

# 檢查現有訂單
existing_orders = []
for row in range(2, ws.max_row + 1):
    order_no = ws.cell(row=row, column=1).value
    if order_no and order_no != 'Ord No.' and str(order_no).strip():
        try:
            existing_orders.append(float(order_no))
        except:
            pass

# 訂單548798
if 548798 not in existing_orders:
    # 尋找第一個空行
    start_row = 2
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value is None or cell_value == '':
            start_row = row
            break
    
    # 訂單548798數據
    order_data = {
        'ord_no': 548798,
        'time': '11:30',
        'trip': 'K11-故宮-K11',
        'driver': 'Eva',
        'plate': 'TZ9800',
        'mobile': '97062468',
        'pick': '11:30',
        'drop': '16:25',
        'extra_fees': '包車',
        'ot': 0,
        'parking': 0,
        'basic_price': 400,
        'driver_fee': 1150,  # 包車價格
        'bonus': 30
    }
    
    # 輸入訂單數據
    row_num = start_row
    ws.cell(row=row_num, column=1, value=order_data['ord_no'])
    ws.cell(row=row_num, column=2, value=order_data['time'])
    ws.cell(row_num, column=3, value=order_data['trip'])
    ws.cell(row=row_num, column=4, value=f'{order_data["driver"]} {order_data["plate"]} {order_data["mobile"]}')
    ws.cell(row=row_num, column=5, value='Done')
    ws.cell(row=row_num, column=6, value=order_data['pick'])
    ws.cell(row=row_num, column=7, value=order_data['drop'])
    ws.cell(row=row_num, column=8, value=order_data['extra_fees'])
    ws.cell(row=row_num, column=9, value=order_data['pick'])
    ws.cell(row=row_num, column=10, value=f'{order_data["driver"]} {order_data["plate"]} {order_data["mobile"]}')
    ws.cell(row=row_num, column=11, value=order_data['ot'])
    ws.cell(row=row_num, column=12, value=order_data['parking'])
    ws.cell(row=row_num, column=13, value=order_data['basic_price'])
    ws.cell(row=row_num, column=14, value=order_data['driver_fee'])
    ws.cell(row=row_num, column=15, value=order_data['bonus'])
    
    print(f'✅ 已輸入訂單548798到第{row_num}行')
else:
    print('⚠️ 訂單548798已存在')

# 處理訂單548797（4月4日）
print('\n=== 處理訂單548797（4月4日）===')
if '4APR2026' in wb.sheetnames:
    ws = wb['4APR2026']
else:
    ws = wb.create_sheet('4APR2026')
    headers = ['Ord No.', 'Time', 'Trip', 'Driver', 'Status', 'Pick', 'Drop', '', '', 'Driver.Details', 'Dvr.OT', 'Dvr.C', 'Basic', 'Dvr.F', 'Bonus']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

# 檢查現有訂單
existing_orders = []
for row in range(2, ws.max_row + 1):
    order_no = ws.cell(row=row, column=1).value
    if order_no and order_no != 'Ord No.' and str(order_no).strip():
        try:
            existing_orders.append(float(order_no))
        except:
            pass

# 訂單548797
if 548797 not in existing_orders:
    # 尋找第一個空行
    start_row = 2
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value is None or cell_value == '':
            start_row = row
            break
    
    # 訂單548797數據
    order_data = {
        'ord_no': 548797,
        'time': '15:00',
        'trip': 'K11-東涌-映灣園-k11',
        'driver': 'Eva',
        'plate': 'TZ9800',
        'mobile': '97062468',
        'pick': '15:00',
        'drop': '18:56',
        'extra_fees': '東涌附加費$200, 停車場$57',
        'ot': 200,  # 東涌附加費
        'parking': 57,
        'basic_price': 400,
        'driver_fee': 657,  # $400 + $200 + $57
        'bonus': 30
    }
    
    # 輸入訂單數據
    row_num = start_row
    ws.cell(row=row_num, column=1, value=order_data['ord_no'])
    ws.cell(row=row_num, column=2, value=order_data['time'])
    ws.cell(row=row_num, column=3, value=order_data['trip'])
    ws.cell(row=row_num, column=4, value=f'{order_data["driver"]} {order_data["plate"]} {order_data["mobile"]}')
    ws.cell(row=row_num, column=5, value='Done')
    ws.cell(row=row_num, column=6, value=order_data['pick'])
    ws.cell(row=row_num, column=7, value=order_data['drop'])
    ws.cell(row=row_num, column=8, value=order_data['extra_fees'])
    ws.cell(row=row_num, column=9, value=order_data['pick'])
    ws.cell(row=row_num, column=10, value=f'{order_data["driver"]} {order_data["plate"]} {order_data["mobile"]}')
    ws.cell(row=row_num, column=11, value=order_data['ot'])
    ws.cell(row_num, column=12, value=order_data['parking'])
    ws.cell(row_num, column=13, value=order_data['basic_price'])
    ws.cell(row_num, column=14, value=order_data['driver_fee'])
    ws.cell(row_num, column=15, value=order_data['bonus'])
    
    print(f'✅ 已輸入訂單548797到第{row_num}行')
else:
    print('⚠️ 訂單548797已存在')

# 保存修改後嘅文件
output_path = '/data/workspace/ACS_Order_APR2026_Final.xlsx'
wb.save(output_path)
print(f'\n📁 文件已保存到: {output_path}')

# 驗證5APR2026 sheet
print('\n=== 5APR2026 Sheet 驗證 ===')
ws_5apr = wb['5APR2026']
print(f'Sheet大小: {ws_5apr.max_row}行 x {ws_5apr.max_column}列')

# 檢查所有訂單
orders_in_5apr = []
for row in range(2, ws_5apr.max_row + 1):
    order_no = ws_5apr.cell(row=row, column=1).value
    if order_no and order_no != 'Ord No.' and str(order_no).strip():
        try:
            orders_in_5apr.append(float(order_no))
        except:
            pass

print(f'5APR2026 現有訂單: {orders_in_5apr}')

print('\n✅ 所有訂單處理完成！')
print('📊 處理摘要:')
print('  - 訂單548798（4月5日）: 包車訂單')
print('  - 訂單548797（4月4日）: 東涌附加費 + 停車場')
print('  - 訂單550414（4月5日）: 兒童座椅配置')
print('  - 準備上傳到GitHub')