import openpyxl
from openpyxl.styles import Font, Alignment
import os

def input_april3_orders():
    try:
        file_path = '/data/workspace/ACS_Order_APR2026_Corrected.xlsx'
        
        # 加載工作簿
        wb = openpyxl.load_workbook(file_path)
        
        # 檢查 3APR2026 工作表是否存在
        if '3APR2026' in wb.sheetnames:
            ws = wb['3APR2026']
            print('使用現有的 3APR2026 工作表')
        else:
            # 創建新的 3APR2026 工作表
            ws = wb.create_sheet('3APR2026')
            print('創建新的 3APR2026 工作表')
            
            # 設置列標題
            headers = ['Ord No.', 'Time', 'Trip', 'Driver', 'Status', 'Pick', 'Drop', 'Decp', 'Ptime', 'Driver.Details', 'Dvr.OT', 'Dvr.C', 'Driver.C+OT', 'Bouns']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
        
        print(f'工作表大小: {ws.max_row}行 x {ws.max_column}列')
        
        # 4月3日訂單數據
        orders_apr3 = [
            {
                'ord_no': '549801',
                'time': '3:30',
                'trip': 'TST > AP',
                'driver': 'Eva TZ9800 97062468',
                'plate': 'TZ9800',
                'mobile': '97062468',
                'pick': '3:38',
                'drop': '4:06',
                'decp': '',
                'ptime': '3:38',
                'driver_details': 'OT$60',
                'ot': '60',
                'status': '',
                'driver_c': '400'
            },
            {
                'ord_no': '549803',
                'time': '5:00',
                'trip': 'TST > AP',
                'driver': 'Ricky 黎 ZJ6071 51089715',
                'plate': 'ZJ6071',
                'mobile': '51089715',
                'pick': '05:00',
                'drop': '05:33',
                'decp': '',
                'ptime': '05:00',
                'driver_details': 'OT$60',
                'ot': '60',
                'status': '',
                'driver_c': '400'
            },
            {
                'ord_no': '549802',
                'time': '3:30',
                'trip': 'MB179 > TST',
                'driver': 'Ricky 黎 ZJ6071 51089715',
                'plate': 'ZJ6071',
                'mobile': '51089715',
                'pick': '04:06',
                'drop': '04:35',
                'decp': '',
                'ptime': '04:06',
                'driver_details': 'OT$60',
                'ot': '60',
                'status': '',
                'driver_c': '400'
            },
            {
                'ord_no': '549802',
                'time': '3:30',
                'trip': 'MB179 > TST',
                'driver': 'Cheuk/卓師傅 CK9898 96695358',
                'plate': 'CK9898',
                'mobile': '96695358',
                'pick': '04:06',
                'drop': '04:35',
                'decp': '',
                'ptime': '04:06',
                'driver_details': 'P$35',
                'ot': '35',
                'status': '',
                'driver_c': '400'
            }
        ]
        
        # 尋找第一個空行
        start_row = 2
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value is None or cell_value == '':
                start_row = row
                break
        
        print(f'從第{start_row}行開始輸入訂單')
        
        # 輸入訂單數據
        for i, order in enumerate(orders_apr3):
            row_num = start_row + i
            
            # 設置單元格數據 - 根據正確的四月份格式
            ws.cell(row=row_num, column=1, value=order['ord_no'])  # Ord No.
            ws.cell(row=row_num, column=2, value=order['time'])    # Time
            
            # 轉換時間格式為 HH:MM:00
            time_str = order['time']
            if ':' in time_str:
                time_parts = time_str.split(':')
                if len(time_parts) == 2:
                    formatted_time = f"{time_parts[0]}:{time_parts[1]}:00"
                else:
                    formatted_time = time_str
            else:
                formatted_time = time_str
            
            ws.cell(row=row_num, column=2, value=formatted_time)  # Time (with seconds)
            ws.cell(row=row_num, column=3, value=order['trip'])    # Trip
            ws.cell(row=row_num, column=4, value=order['driver'])  # Driver (完整信息)
            ws.cell(row=row_num, column=5, value=order['status'])  # Status
            ws.cell(row=row_num, column=6, value=order['pick'])    # Pick
            
            # 轉換上車時間格式
            pick_str = order['pick']
            if ':' in pick_str:
                pick_parts = pick_str.split(':')
                if len(pick_parts) == 2:
                    formatted_pick = f"{pick_parts[0]}:{pick_parts[1]}:00"
                else:
                    formatted_pick = pick_str
            else:
                formatted_pick = pick_str
            
            ws.cell(row=row_num, column=6, value=formatted_pick)  # Pick (with seconds)
            ws.cell(row=row_num, column=7, value=order['drop'])    # Drop
            
            # 轉換下車時間格式
            drop_str = order['drop']
            if ':' in drop_str:
                drop_parts = drop_str.split(':')
                if len(drop_parts) == 2:
                    formatted_drop = f"{drop_parts[0]}:{drop_parts[1]}:00"
                else:
                    formatted_drop = drop_str
            else:
                formatted_drop = drop_str
            
            ws.cell(row=row_num, column=7, value=formatted_drop)  # Drop (with seconds)
            ws.cell(row=row_num, column=8, value=order['decp'])    # Decp (空)
            ws.cell(row=row_num, column=9, value=formatted_pick)   # Ptime (與 Pick 相同)
            ws.cell(row=row_num, column=10, value=order['driver_details'])  # Driver.Details (備註)
            
            # 設置 OT 數值
            if order['ot']:
                try:
                    ws.cell(row=row_num, column=11, value=float(order['ot']))  # Dvr.OT
                except ValueError:
                    ws.cell(row=row_num, column=11, value=0)  # Dvr.OT
            
            # 設置 Dvr.C (基本費)
            if order['driver_c']:
                try:
                    ws.cell(row=row_num, column=12, value=float(order['driver_c']))  # Dvr.C
                except ValueError:
                    ws.cell(row=row_num, column=12, value=400)  # Dvr.C
            
            # 設置 Bouns (除 no show 外都為 30)
            ws.cell(row=row_num, column=14, value=30)  # Bouns
            
            print(f'已輸入第{i+1}筆訂單到第{row_num}行: {order["ord_no"]} - {order["trip"]}')
        
        # 保存修改後的文件
        output_path = '/data/workspace/ACS_Order_APR2026_With_April3.xlsx'
        wb.save(output_path)
        print(f'文件已保存到: {output_path}')
        
        # 返回輸入的訂單數量
        return len(orders_apr3)
        
    except Exception as e:
        print(f'處理文件時出錯: {e}')
        import traceback
        traceback.print_exc()
        return 0

if __name__ == '__main__':
    orders_count = input_april3_orders()
    print(f'成功輸入 {orders_count} 筆訂單到 3APR2026 工作表')