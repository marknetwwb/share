import openpyxl

# 讀取最新嘅Excel文件
file_path = '/data/workspace/ACS_Order_APR2026_CS---997ab6ed-a3ce-4c12-bc31-052ab333e6c6.xlsx'
wb = openpyxl.load_workbook(file_path)

print('=== 開始處理所有訂單 ===')

def process_order(order_data, target_date):
    # 處理訂單數據
    if target_date in wb.sheetnames:
        ws = wb[target_date]
    else:
        ws = wb.create_sheet(target_date)
        headers = ['Ord No.', 'Time', 'Trip', 'Driver', 'Status', 'Pick', 'Drop', '', '', 'Driver.Details', 'Dvr.OT', 'Dvr.C', 'Basic', 'Dvr.F', 'Bonus']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
    
    # 檢查現有訂單
    existing_orders = []
    for row in range(2, ws.max_row + 1):
        order_no = ws.cell(row=row, column=1).value
        if order_no and order_no != 'Ord No.' and str(order_no).strip():
            try:
                existing_orders.append(float(order_no))
            except:
                pass
    
    print(f'{target_date} 現有訂單: {existing_orders}')
    
    # 檢查訂單是否已存在
    if order_data['ord_no'] in existing_orders:
        print(f'訂單 {order_data["ord_no"]} 已存在，跳過')
        return
    
    # 尋找第一個空行
    start_row = 2
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value is None or cell_value == '':
            start_row = row
            break
    
    print(f'從第{start_row}行開始輸入訂單 {order_data["ord_no"]}')
    
    # 輸入訂單數據
    if 'cars' in order_data:  # 多台車
        for i, car in enumerate(order_data['cars']):
            row_num = start_row + i
            
            ws.cell(row=row_num, column=1, value=order_data['ord_no'])
            ws.cell(row=row_num, column=2, value=order_data['time'])
            ws.cell(row=row_num, column=3, value=order_data['trip'])
            ws.cell(row=row_num, column=4, value=f'{car["driver"]} {car["plate"]} {car["mobile"]}')
            ws.cell(row=row_num, column=5, value='Done')
            ws.cell(row=row_num, column=6, value=car['pick'])
            ws.cell(row=row_num, column=7, value=car['drop'])
            ws.cell(row=row_num, column=8, value=car['extra_fees'])
            ws.cell(row=row_num, column=9, value=car['pick'])
            ws.cell(row_num, column=10, value=f'{car["driver"]} {car["plate"]} {car["mobile"]}')
            ws.cell(row=row_num, column=11, value=car['ot'])
            ws.cell(row=row_num, column=12, value=car['parking'])
            ws.cell(row_num, column=13, value=car['basic_price'])
            ws.cell(row=row_num, column=14, value=car['driver_fee'])
            ws.cell(row=row_num, column=15, value=car['bonus'])
            
            print(f'已輸入第{i+1}台車到第{row_num}行: {car["driver"]} {car["plate"]}')
    else:  # 單台車
        row_num = start_row
        
        ws.cell(row=row_num, column=1, value=order_data['ord_no'])
        ws.cell(row=row_num, column=2, value=order_data['time'])
        ws.cell(row=row_num, column=3, value=order_data['trip'])
        ws.cell(row=row_num, column=4, value=f'{order_data["driver"]} {order_data["plate"]} {order_data["mobile"]}')
        ws.cell(row=row_num, column=5, value='Done')
        ws.cell(row=row_num, column=6, value=order_data['pick'])
        ws.cell(row=row_num, column=7, value=order_data['drop'])
        ws.cell(row=row_num, column=8, value=order_data['extra_fees'])
        ws.cell(row=row_num, column=9, value=order_data['pick'])
        ws.cell(row=row_num, column=10, value=f'{order_data["driver"]} {order_data["plate"]} {order_data["mobile"]}')
        ws.cell(row=row_num, column=11, value=order_data['ot'])
        ws.cell(row=row_num, column=12, value=order_data['parking'])
        ws.cell(row=row_num, column=13, value=order_data['basic_price'])
        ws.cell(row=row_num, column=14, value=order_data['driver_fee'])
        ws.cell(row=row_num, column=15, value=order_data['bonus'])
        
        print(f'已輸入訂單 {order_data["ord_no"]} 到第{row_num}行')

# 處理訂單549905（4月4日）
print('\n=== 處理訂單549905（4月4日）===')
order_549905 = {
    'ord_no': 549905,
    'time': '00:15',
    'trip': 'TST > Ap',
    'driver': 'Eva',
    'plate': 'TZ9800',
    'mobile': '97062468',
    'pick': '00:43',
    'drop': '01:16',
    'extra_fees': 'OT$60',
    'ot': 60,
    'parking': 0,
    'basic_price': 400,
    'driver_fee': 460,
    'bonus': 30
}
process_order(order_549905, '4APR2026')

# 處理訂單548797（4月4日）
print('\n=== 處理訂單548797（4月4日）===')
order_548797 = {
    'ord_no': 548797,
    'time': '15:00',
    'trip': 'K11-東涌-映灣園-k11',
    'driver': 'Eva',
    'plate': 'TZ9800',
    'mobile': '97062468',
    'pick': '15:00',
    'drop': '18:56',
    'extra_fees': '東涌附加費$200, 停車場$57',
    'ot': 200,
    'parking': 57,
    'basic_price': 400,
    'driver_fee': 657,
    'bonus': 30
}
process_order(order_548797, '4APR2026')

# 處理訂單546271（3月11日）
print('\n=== 處理訂單546271（3月11日）===')
order_546271 = {
    'ord_no': 546271,
    'time': '12:06',
    'trip': 'Ap - yau ma tei - K11 - che kung - Disney',
    'driver': 'Keith',
    'plate': '',
    'mobile': '',
    'pick': '12:06',
    'drop': '16:52',
    'extra_fees': '大嶼山附加費2X200=$400, 隧道費$16, 停車場$16',
    'ot': 400,
    'parking': 16,
    'basic_price': 400,
    'driver_fee': 832,
    'bonus': 30
}
process_order(order_546271, '11MAR2026')

# 處理訂單549906（4月4日）
print('\n=== 處理訂單549906（4月4日）===')
order_549906 = {
    'ord_no': 549906,
    'time': '1:15',
    'trip': 'MB300 > TST',
    'cars': [
        {
            'driver': 'Eva',
            'plate': 'TZ9800',
            'mobile': '97062468',
            'pick': '1:58',
            'drop': '2:28',
            'extra_fees': 'OT$60, P$35',
            'ot': 60,
            'parking': 0,
            'basic_price': 400,
            'driver_fee': 460,
            'bonus': 30
        },
        {
            'driver': 'Soo',
            'plate': 'FN655',
            'mobile': '60977620',
            'pick': '1:58',
            'drop': '2:28',
            'extra_fees': 'OT$60, P$35',
            'ot': 60,
            'parking': 35,
            'basic_price': 400,
            'driver_fee': 495,
            'bonus': 30
        }
    ]
}
process_order(order_549906, '4APR2026')

# 處理訂單550414（5月4日）
print('\n=== 處理訂單550414（5月4日）===')
order_550414 = {
    'ord_no': 550414,
    'time': '20:55',
    'trip': 'CX581 > Repulse Bay Gardens',
    'cars': [
        {
            'driver': 'Eva',
            'plate': 'TZ9800',
            'mobile': '97062468',
            'pick': '22:01',
            'drop': '22:50',
            'extra_fees': '南區$60, Bb seat x2',
            'ot': 60,
            'parking': 0,
            'basic_price': 400,
            'driver_fee': 720,
            'bonus': 30
        },
        {
            'driver': 'Ricky 黎',
            'plate': 'ZJ6071',
            'mobile': '51089715',
            'pick': '22:01',
            'drop': '22:50',
            'extra_fees': '南區$60, Bb seat x1',
            'ot': 60,
            'parking': 0,
            'basic_price': 400,
            'driver_fee': 640,
            'bonus': 30
        }
    ]
}
process_order(order_550414, '5APR2026')

# 保存修改後的文件
output_path = '/data/workspace/ACS_Order_APR2026_Corrected.xlsx'
wb.save(output_path)
print(f'\n📁 文件已保存到: {output_path}')

# 驗證最終結果
print('\n=== 最終驗證 ===')
for sheet_name in ['4APR2026', '5APR2026', '11MAR2026']:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        orders_count = 0
        order_numbers = []
        for row in range(2, ws.max_row + 1):
            order_no = ws.cell(row=row, column=1).value
            if order_no and order_no != 'Ord No.' and str(order_no).strip():
                orders_count += 1
                order_numbers.append(order_no)
        print(f'{sheet_name}: {orders_count} 個訂單 - {order_numbers}')

print('\n✅ 所有訂單處理完成！')
"