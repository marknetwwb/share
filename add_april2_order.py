import openpyxl

def add_april2_order():
    try:
        file_path = '/data/workspace/ACS_Order_APR2026_With_April3.xlsx'
        
        # 加載工作簿
        wb = openpyxl.load_workbook(file_path)
        ws = wb['2APR2026']
        
        print(f'2APR2026 工作表大小: {ws.max_row}行 x {ws.max_column}列')
        
        # 4月2日訂單 #549349
        order_apr2 = {
            'ord_no': '549349',
            'time': '6:15',
            'trip': 'CX234 > hotel icon',
            'driver': '陳師傅（豪） WH9918 5575 5138',
            'plate': 'WH9918',
            'mobile': '5575 5138',
            'pick': '07:30',
            'drop': '08:00',
            'decp': '',
            'ptime': '07:30',
            'driver_details': '公司 waiting 0.5h 早',
            'ot': '',
            'status': '',
            'driver_c': '400'
        }
        
        # 尋找第一個空行
        start_row = 2
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value is None or cell_value == '':
                start_row = row
                break
        
        print(f'在第{start_row}行輸入訂單 #549349')
        
        # 輸入訂單數據 - 根據正確的四月份格式
        ws.cell(row=start_row, column=1, value=order_apr2['ord_no'])  # Ord No.
        
        # 轉換時間格式為 HH:MM:00
        time_str = order_apr2['time']
        if ':' in time_str:
            time_parts = time_str.split(':')
            if len(time_parts) == 2:
                formatted_time = f"{time_parts[0]}:{time_parts[1]}:00"
            else:
                formatted_time = time_str
        else:
            formatted_time = time_str
        
        ws.cell(row=start_row, column=2, value=formatted_time)  # Time (with seconds)
        ws.cell(row=start_row, column=3, value=order_apr2['trip'])  # Trip
        ws.cell(row=start_row, column=4, value=order_apr2['driver'])  # Driver (完整信息)
        ws.cell(row=start_row, column=5, value=order_apr2['status'])  # Status
        
        # 轉換上車時間格式
        pick_str = order_apr2['pick']
        if ':' in pick_str:
            pick_parts = pick_str.split(':')
            if len(pick_parts) == 2:
                formatted_pick = f"{pick_parts[0]}:{pick_parts[1]}:00"
            else:
                formatted_pick = pick_str
        else:
            formatted_pick = pick_str
        
        ws.cell(row=start_row, column=6, value=formatted_pick)  # Pick (with seconds)
        
        # 轉換下車時間格式
        drop_str = order_apr2['drop']
        if ':' in drop_str:
            drop_parts = drop_str.split(':')
            if len(drop_parts) == 2:
                formatted_drop = f"{drop_parts[0]}:{drop_parts[1]}:00"
            else:
                formatted_drop = drop_str
        else:
            formatted_drop = drop_str
        
        ws.cell(row=start_row, column=7, value=formatted_drop)  # Drop (with seconds)
        ws.cell(row=start_row, column=8, value=order_apr2['decp'])  # Decp (空)
        ws.cell(row=start_row, column=9, value=formatted_pick)  # Ptime (與 Pick 相同)
        ws.cell(row=start_row, column=10, value=order_apr2['driver_details'])  # Driver.Details (備註)
        
        # 設置 Dvr.C (基本費)
        if order_apr2['driver_c']:
            try:
                ws.cell(row=start_row, column=12, value=float(order_apr2['driver_c']))  # Dvr.C
            except ValueError:
                ws.cell(row=start_row, column=12, value=400)  # Dvr.C
        
        # 設置 Bouns (除 no show 外都為 30)
        ws.cell(row=start_row, column=14, value=30)  # Bouns
        
        # 設置 Driver.C+OT 公式
        ws.cell(row=start_row, column=13, value=f'=K{start_row}+L{start_row}')  # Driver.C+OT
        
        print(f'已成功輸入訂單 #549349 到第{start_row}行')
        order_content = f'{order_apr2["ord_no"]} - {order_apr2["trip"]}'
        print(f'訂單內容: {order_content}')
        
        # 保存修改後的文件
        output_path = '/data/workspace/ACS_Order_APR2026_With_April2_Corrected.xlsx'
        wb.save(output_path)
        print(f'文件已保存到: {output_path}')
        
        return True
        
    except Exception as e:
        print(f'處理文件時出錯: {e}')
        import traceback
        traceback.print_exc()
        return False

if __name__ == '__main__':
    success = add_april2_order()
    if success:
        print('✅ 成功將 4月2 日訂單 #549349 添加到 Excel 檔案')
    else:
        print('❌ 添加失敗')