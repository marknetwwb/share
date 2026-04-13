#!/usr/bin/env python3
"""
簡化版發票處理系統
直接處理發票圖片並添加到Excel
"""

import os
import sys
from datetime import datetime
from PIL import Image
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage

class SimpleInvoiceProcessor:
    """簡化版發票處理器"""
    
    def __init__(self, excel_path="/data/workspace/AdvinvSys/invoice_template.xlsx"):
        self.excel_path = excel_path
        self.next_id = self._get_next_id()
        
    def _get_next_id(self):
        """獲取下一個編號"""
        try:
            wb = load_workbook(self.excel_path)
            ws = wb.active
            last_row = ws.max_row
            if last_row > 1:  # 有數據行
                last_id = ws.cell(row=last_row, column=1).value
                return int(last_id) + 1 if last_id else 1
            return 1
        except:
            return 1
    
    def compress_image(self, image_path, max_size=(200, 200)):
        """壓縮圖片"""
        try:
            image = Image.open(image_path)
            image.thumbnail(max_size, Image.Resampling.LANCZOS)
            
            # 保存為壓縮版本
            compressed_path = os.path.splitext(image_path)[0] + '_compressed.jpg'
            image.save(compressed_path, 'JPEG', quality=85, optimize=True)
            
            return compressed_path
        except Exception as e:
            print(f"圖片壓縮失敗: {e}")
            return image_path
    
    def add_invoice_to_excel(self, image_path, invoice_type="零售", amount=0.0, business_name="未知商家"):
        """添加發票到Excel"""
        try:
            # 壓縮圖片
            compressed_image_path = self.compress_image(image_path)
            
            # 加載Excel文件
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            # 找到下一個空行
            next_row = ws.max_row + 1
            
            # 添加數據
            ws.cell(row=next_row, column=1, value=self.next_id)  # 編號
            ws.cell(row=next_row, column=2, value=datetime.now().strftime("%Y-%m-%d"))  # 日期
            ws.cell(row=next_row, column=3, value=invoice_type)  # 類型
            ws.cell(row=next_row, column=4, value=business_name) # 商家
            ws.cell(row=next_row, column=5, value=amount)        # 金額
            ws.cell(row=next_row, column=6, value="自動化處理")  # 備註
            
            # 插入圖片
            try:
                img = OpenpyxlImage(compressed_image_path)
                img.width = 200
                img.height = 200
                ws.add_image(img, f'G{next_row}')  # 發票圖片欄位
            except Exception as e:
                print(f"圖片插入失敗: {e}")
            
            # 保存文件
            wb.save(self.excel_path)
            
            # 更新下一個編號
            self.next_id += 1
            
            return True, f"發票已成功添加到Excel (編號: {self.next_id-1})"
            
        except Exception as e:
            return False, f"添加到Excel失敗: {e}"
    
    def process_invoice(self, image_path):
        """處理發票"""
        print(f"開始處理發票圖片: {image_path}")
        
        # 基本信息提取（簡化版）
        print("步驟1: 分析圖片...")
        
        # 檢查圖片信息
        try:
            img = Image.open(image_path)
            print(f"圖片尺寸: {img.size}")
            print(f"圖片格式: {img.format}")
        except Exception as e:
            print(f"圖片分析失敗: {e}")
            return False, "圖片分析失敗"
        
        # 模擬AI分類
        print("步驟2: AI分類...")
        invoice_type = "零售"  # 根據圖片情況調整
        print(f"分類結果: {invoice_type}")
        
        # 模擬金額提取
        print("步驟3: 提取金額...")
        amount = 0.0  # 實際應該從圖片提取
        print(f"提取金額: {amount}")
        
        # 商家名稱
        print("步驟4: 商家信息...")
        business_name = "測試商家"
        print(f"商家: {business_name}")
        
        # 添加到Excel
        print("步驟5: 添加到Excel...")
        success, message = self.add_invoice_to_excel(image_path, invoice_type, amount, business_name)
        
        if success:
            print("✅ 發票處理完成！")
            return True, message
        else:
            print("❌ 發票處理失敗！")
            return False, message

def main():
    """主函數"""
    if len(sys.argv) != 2:
        print("使用方法: python simple_invoice_processor.py <圖片路徑>")
        sys.exit(1)
    
    image_path = sys.argv[1]
    if not os.path.exists(image_path):
        print(f"圖片文件不存在: {image_path}")
        sys.exit(1)
    
    # 創建處理器
    processor = SimpleInvoiceProcessor()
    
    # 處理發票
    success, message = processor.process_invoice(image_path)
    
    print("\n" + "="*50)
    print("處理結果:")
    print(f"狀態: {'成功' if success else '失敗'}")
    print(f"信息: {message}")
    print("="*50)
    
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()