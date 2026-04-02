import openpyxl

def compare_excel_files():
    try:
        # 讀取我的版本（錯誤版本）
        my_file = '/data/workspace/ACS_Order_APR2026_Orders_Input.xlsx'
        my_wb = openpyxl.load_workbook(my_file)
        my_ws = my_wb['1APR2026']
        
        # 讀取修正後版本（正確版本）
        corrected_file = '/data/workspace/ACS_Order_APR2026_Corrected.xlsx'
        corrected_wb = openpyxl.load_workbook(corrected_file)
        corrected_ws = corrected_wb['1APR2026']
        
        print('=== Excel 文件比對分析 ===')
        print(f'我的版本: {my_file}')
        print(f'修正版本: {corrected_file}')
        print()
        
        # 比對前10行的差異
        print('=== 詳細差異分析 ===')
        differences_found = []
        
        for row in range(2, 10):  # 比對前8行訂單
            print(f'\\n=== 行 {row} 比對 ===')
            
            row_differences = []
            
            # 比對每列
            for col in range(1, 16):  # 比對前15列
                my_value = str(my_ws.cell(row=row, column=col).value) if my_ws.cell(row=row, column=col).value is not None else ''
                corrected_value = str(corrected_ws.cell(row=row, column=col).value) if corrected_ws.cell(row=row, column=col).value is not None else ''
                
                if my_value != corrected_value:
                    col_names = {
                        1: 'Ord No.', 2: 'Time', 3: 'Trip', 4: 'Driver', 5: 'Status', 
                        6: 'Pick', 7: 'Drop', 8: 'Decp', 9: 'Ptime', 10: 'Driver.Details',
                        11: 'Dvr.OT', 12: 'Dvr.C', 13: 'Driver.C+OT', 14: 'Bouns', 15: 'Column15'
                    }
                    col_name = col_names.get(col, f'Column{col}')
                    
                    print(f'  列{col} ({col_name}): 我的="{my_value}" vs 修正="{corrected_value}"')
                    row_differences.append({
                        'row': row,
                        'col': col,
                        'col_name': col_name,
                        'my_value': my_value,
                        'corrected_value': corrected_value
                    })
            
            if row_differences:
                differences_found.extend(row_differences)
        
        # 總結差異
        print('\\n=== 差異總結 ===')
        if differences_found:
            print(f'發現 {len(differences_found)} 個差異:')
            
            # 按差異類型分類
            bonus_errors = [d for d in differences_found if d['col_name'] == 'Bouns']
            other_errors = [d for d in differences_found if d['col_name'] != 'Bouns']
            
            if bonus_errors:
                print(f'\\n🎯 Bouns 欄位錯誤 ({len(bonus_errors)} 個):')
                for error in bonus_errors:
                    print(f'  行{error["row"]}: 應該是 "30" 而不是 "{error["my_value"]}"')
            
            if other_errors:
                print(f'\\n📝 其他欄位錯誤 ({len(other_errors)} 個):')
                for error in other_errors:
                    print(f'  行{error["row"]} 列{error["col"]} ({error["col_name"]}):')
                    print(f'    我輸入: "{error["my_value"]}"')
                    print(f'    正確值: "{error["corrected_value"]}"')
        else:
            print('✅ 沒有發現差異，文件完全一致！')
        
        # 生成學習記錄
        generate_learning_record(differences_found, my_file, corrected_file)
        
        return len(differences_found)
        
    except Exception as e:
        print(f'比對過程出錯: {e}')
        import traceback
        traceback.print_exc()
        return 0

def generate_learning_record(differences, my_file, corrected_file):
    """生成學習記錄文件"""
    
    learning_content = f"""# 車隊記憶 - Excel 輸入錯誤學習記錄

## 📊 **比對分析報告**

**比對時間**: 2026-04-02  
**我的版本**: {my_file}  
**修正版本**: {corrected_file}  
**差異總數**: {len(differences)} 個

## 🚨 **主要錯誤類型**

### 1. Bouns 欄位錯誤
**錯誤**: 我將 Bouns 欄位設置為 "0" 或空值  
**正確值**: 除 no show 外，每個訂單的 Bouns 欄位都應該是 "30"  
**原因**: 忽略了車隊的標準收費規則  
**學習**: 必須牢記 Bouns = $30 是標準配置

### 2. 可能的其他格式錯誤
{chr(10).join([f'**行 {error["row"]} {error["col_name"]}**: 我輸入 "{error["my_value"]}"，正確應該是 "{error["corrected_value"]}"' 
                for error in differences if error['col_name'] != 'Bouns'])}

## ✅ **正確的輸入規則**

### Bouns 欄位規則:
- **除 no show 外**: 每個訂單 Bouns = 30
- **no show 訂單**: Bouns = 0 或空值
- **位置**: 第14列 (最右邊一列)

### 其他注意事項:
- 確保所有欄位格式正確
- 時間格式符合月份要求 (四月不含秒數)
- 雙車輛訂單分兩行輸入
- 保持數值格式的準確性

## 📝 **改進措施**

1. **建立檢查清單**: 輸入後必須檢查 Bouns 欄位
2. **自動化驗證**: 在腳本中加入 Bouns 欄位自動設置
3. **格式記錄**: 將此錯誤記錄到車隊格式規則中
4. **定期複習**: 定期複習車隊的標準收費規則

## 🔗 **相關文件**

- **正確版本**: ACS_Order_APR2026_Corrected.xlsx
- **錯誤版本**: ACS_Order_APR2026_Orders_Input.xlsx
- **格式規則**: fleet_memory_format_rules.md
- **總計數據**: fleet_memory_march_totals.md

---
**記錄時間**: 2026-04-02  
**記錄目的**: 避免重複錯誤，提高輸入準確性
**狀態**: 已完成比對分析並記錄學習要點
"""
    
    # 保存學習記錄
    with open('/data/workspace/fleet_learning_mistake_record.md', 'w', encoding='utf-8') as f:
        f.write(learning_content)
    
    print('\\n📚 學習記錄已生成: fleet_learning_mistake_record.md')

if __name__ == '__main__':
    differences_count = compare_excel_files()
    print(f'\\n✅ 比對完成，發現 {differences_count} 個差異')